import statistics
import matplotlib.pyplot as plt
import openpyxl


def read(id):
    data_collect=[]
    filename=id+".txt"
    file_path = filename  # 替换为你的文件路径
    with open(file_path, 'r') as file:
        for line in file:
            line = line.rstrip('\n')
            data=line.split(" ")
            data_collect.append(float(data[1]))
    return data_collect
            
x=[i+0.5 for i in range(500)]
name=["1e2","3e2","1e3","3e3","1e4","3e4","1e5"]
color=['blue','green','red','purple','orange','cyan','magenta']

workbook = openpyxl.Workbook()
sheet = workbook.active

for i in range(7):
    p=read(name[i])[:500]
    for j in range(500):
        sheet.cell(row=j+1, column=i+2, value=p[j])
for j in range(500):
    sheet.cell(row=j+1, column=1, value=j/500)


workbook.save("output.xlsx")
workbook.close()   

